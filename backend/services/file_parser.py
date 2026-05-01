import hashlib
import logging
from io import BytesIO
from pathlib import Path
from typing import Optional

import pdfplumber
from docx import Document
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

SUPPORTED_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".txt": "text/plain",
    ".md": "text/markdown",
}


class FileParserError(Exception):
    pass


class FileParser:
    @staticmethod
    def parse(file_content: bytes, filename: str) -> dict:
        suffix = Path(filename).suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            raise FileParserError(f"不支持的文件格式: {suffix}")

        if len(file_content) > MAX_FILE_SIZE:
            raise FileParserError(f"文件大小超过限制 ({MAX_FILE_SIZE // 1024 // 1024}MB)")

        content = ""
        if suffix == ".pdf":
            content = FileParser._parse_pdf(file_content)
        elif suffix == ".docx":
            content = FileParser._parse_docx(file_content)
        elif suffix == ".xlsx":
            content = FileParser._parse_xlsx(file_content)
        elif suffix in (".txt", ".md"):
            content = FileParser._parse_text(file_content)

        if not content.strip():
            raise FileParserError("文件内容为空")

        return {
            "title": filename,
            "content": content,
            "source_type": f"file_{suffix[1:]}",
            "file_hash": hashlib.md5(file_content).hexdigest(),
        }

    @staticmethod
    def _parse_pdf(file_content: bytes) -> str:
        try:
            with pdfplumber.open(BytesIO(file_content)) as pdf:
                pages = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        pages.append(text)
                return "\n\n".join(pages)
        except Exception as e:
            raise FileParserError(f"PDF 解析失败: {e}")

    @staticmethod
    def _parse_docx(file_content: bytes) -> str:
        try:
            doc = Document(BytesIO(file_content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)
        except Exception as e:
            raise FileParserError(f"Word 解析失败: {e}")

    @staticmethod
    def _parse_xlsx(file_content: bytes) -> str:
        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            sheets_content = []
            for sheet in wb.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    rows.append("\t".join(cells))
                if rows:
                    sheets_content.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
            return "\n\n".join(sheets_content)
        except Exception as e:
            raise FileParserError(f"Excel 解析失败: {e}")

    @staticmethod
    def _parse_text(file_content: bytes) -> str:
        try:
            for encoding in ["utf-8", "gbk", "gb2312", "latin-1"]:
                try:
                    return file_content.decode(encoding)
                except UnicodeDecodeError:
                    continue
            raise FileParserError("无法识别文件编码")
        except Exception as e:
            raise FileParserError(f"文本解析失败: {e}")
